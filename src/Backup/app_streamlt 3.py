# src/app_streamlit.py
# Versión limpia para "Tutor IA para Profesores"
# - Lee Excel o Google Sheet
# - Filtra por materia/año/unidad
# - No muestra fórmulas (celdas que empiezan con '=')
# - Avatar simple (subida + presets)
# - Guardado seguro en hoja 'Consejos'
# Reemplaza tu archivo actual por este contenido.

import os
import json
import datetime
import tempfile
from pathlib import Path
from io import BytesIO

import streamlit as st
import pandas as pd
import openpyxl

# Google Sheets
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_dataframe import get_as_dataframe

# Config
st.set_page_config(page_title="Tutor IA para Profesores", page_icon="🎓", layout="wide")
st.title("🎓 Tutor IA para Profesores")
st.write("Bienvenido/a. Este asistente ayuda a consultar el ecosistema educativo y propone mejoras prácticas (no muestra fórmulas).")

# Cargar OPENAI_KEY (opcional)
OPENAI_KEY = os.getenv("OPENAI_API_KEY")  # si está en .env será usada; si no, app funciona en modo solo-fuentes

# Rutas
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
# Nombre por defecto (si existe)
default_excel = next(DATA_DIR.glob("*.xlsx"), None)
EXCEL_DEFAULT_PATH = str(default_excel) if default_excel else str(DATA_DIR / "Base de Datos Ecosistema Secundaria Aprende.xlsx")

# ------------------------------------------------------------------------
# Utilidades
# ------------------------------------------------------------------------
def sanitize_cell_value(v):
    """Quitar/formatear celdas con fórmulas u otros contenidos no relevantes.
       Si el valor es texto y empieza con '=' lo consideramos fórmula y devolvemos ''.
    """
    try:
        if isinstance(v, str):
            v_strip = v.strip()
            if v_strip.startswith("="):
                return ""  # ocultamos fórmulas
            return v
        # otros valores (num, float, bool, datetime) los devolvemos tal cual
        return v
    except Exception:
        return ""

def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica sanitize_cell_value a todas las celdas y rellena NaN con ''"""
    df = df.copy()
    # convertimos todo a objetos y aplicamos
    return df.fillna("").applymap(sanitize_cell_value)

def extract_candidates_from_sheets(sheets: dict):
    subjects, years, units = set(), set(), set()
    for name, df in sheets.items():
        if df is None or df.empty:
            continue
        for col in df.columns:
            cname = str(col).lower()
            try:
                col_values = df[col].dropna().astype(str).tolist()
            except Exception:
                continue
            if any(k in cname for k in ["materia","asign","subject","disciplina","curso","materi"]):
                subjects.update(col_values)
            if any(k in cname for k in ["año","anio","year","curso","grado","nivel"]):
                years.update(col_values)
            if any(k in cname for k in ["unidad","unidad temática","unidad_t","tema","capítulo","capitulo"]):
                units.update(col_values)
    return sorted([s for s in subjects if s]), sorted([y for y in years if y]), sorted([u for u in units if u])

def show_avatar_bytes(bts, width=140):
    if not bts:
        return
    try:
        txt = bts.decode("utf-8")
        if txt.lstrip().startswith("<svg"):
            st.image("data:image/svg+xml;utf8," + txt, width=width)
            return
    except Exception:
        pass
    st.image(bts, width=width)

# ------------------------------------------------------------------------
# Carga de datos: Google Sheet opcional o Excel local
# ------------------------------------------------------------------------
# ---------- FLOW: Solicitar permiso al propietario por e-mail (one-time token) ----------
import uuid
import json
import urllib.parse
from pathlib import Path
from datetime import datetime

ACCESS_REQS = Path("data/access_requests.json")
ACCESS_GRANTS = Path("data/access_grants.json")
ACCESS_REQS.parent.mkdir(exist_ok=True, parents=True)

def load_json_safe(p):
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def save_json_safe(p, obj):
    p.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

# cargar almacenamiento local
access_reqs = load_json_safe(ACCESS_REQS)
access_grants = load_json_safe(ACCESS_GRANTS)

st.sidebar.markdown("## Acceso al Google Sheet (solicitar permiso)")
requester_email = st.sidebar.text_input("Tu correo (quien solicita acceso)", value=requester_email if 'requester_email' in globals() else "")
sheet_url = st.sidebar.text_input("URL de la hoja de cálculo de Google (opcional)", value=sheet_url if 'sheet_url' in globals() else "")

# intentar extraer email de la service account si se subió el JSON
service_account_email = None
if sa_path and os.path.exists(sa_path):
    try:
        sa_json = json.loads(open(sa_path, "r", encoding="utf-8").read())
        service_account_email = sa_json.get("client_email")
    except Exception:
        service_account_email = None

st.sidebar.markdown("---")
st.sidebar.markdown("### Pedir permiso al propietario (rápido)")

# Generar token único sólo si hay URL y correo solicitante
if st.sidebar.button("Generar solicitud de acceso (crear token)", key="gen_token_btn"):
    if not requester_email or not sheet_url:
        st.sidebar.error("Completá tu correo y la URL del Sheet antes de generar la solicitud.")
    else:
        # crear token
        token = uuid.uuid4().hex[:8].upper()
        timestamp = datetime.utcnow().isoformat() + "Z"
        key = f"{requester_email}|{sheet_url}"
        access_reqs[key] = {"token": token, "requester": requester_email, "sheet_url": sheet_url, "sa_email": service_account_email, "ts": timestamp}
        save_json_safe(ACCESS_REQS, access_reqs)
        st.sidebar.success("Token generado. Revisá el cuerpo del correo para enviarlo al propietario.")
        # preparar body del mail
        sa_line = f"\n\nPor favor compartan el documento con la cuenta de servicio: {service_account_email}\n" if service_account_email else "\n\n(Adjunten la cuenta de servicio o compartan con la cuenta de la app)\n"
        subject = f"Solicitud de acceso a Google Sheet para Tutor IA - {requester_email}"
        body = (
            f"Hola,\n\n"
            f"Solicito acceso al Google Sheet para que el docente {requester_email} pueda usar la aplicación Tutor IA.\n\n"
            f"URL del Sheet:\n{sheet_url}\n"
            f"{sa_line}"
            f"Para confirmar la autorización y que la aplicación lo registre automáticamente, por favor responda con este CÓDIGO (o pégalo en la app):\n\nTOKEN: {token}\n\n"
            "Gracias.\n"
        )
        mailto = f"mailto:?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        st.sidebar.markdown(f"[Abrir cliente de correo para enviar la solicitud]({mailto})", unsafe_allow_html=True)
        st.sidebar.markdown("O copiá el cuerpo del correo y pegalo en tu cliente de correo.")
        st.sidebar.code(body, language="text")

st.sidebar.markdown("---")
st.sidebar.markdown("### Confirmación de autorización (pegar token)")
entered_token = st.sidebar.text_input("Pega aquí el TOKEN que te dio el propietario (si te lo envió)", value="")
if st.sidebar.button("Confirmar token", key="confirm_token_btn"):
    if not entered_token:
        st.sidebar.error("Ingresá el token que te envió el propietario.")
    else:
        # buscar token en access_reqs
        found = None
        for k,v in access_reqs.items():
            if v.get("token") == entered_token.strip():
                found = (k,v)
                break
        if not found:
            st.sidebar.error("Token no encontrado o inválido. Verificá el código.")
        else:
            key, rec = found
            # registrar grant localmente
            grant_key = key  # same key: requester_email|sheet_url
            access_grants[grant_key] = {"requester": rec["requester"], "sheet_url": rec["sheet_url"], "sa_email": rec.get("sa_email"), "granted_ts": datetime.utcnow().isoformat() + "Z"}
            save_json_safe(ACCESS_GRANTS, access_grants)
            # opcional: borrar la solicitud (one-time)
            try:
                del access_reqs[key]
            except KeyError:
                pass
            save_json_safe(ACCESS_REQS, access_reqs)
            st.sidebar.success("Autorización confirmada localmente. Ahora la app la marcará como permitida para este usuario y Sheet.")

# Función utilitaria para verificar si un requester+sheet está autorizado
def is_authorized_locally(requester_email_val, sheet_url_val):
    k = f"{requester_email_val}|{sheet_url_val}"
    return k in access_grants

# Más abajo, antes de intentar leer el Google Sheet, comprobá:
# if sheet_url and sa_path and not is_authorized_locally(requester_email, sheet_url):
#     st.sidebar.warning("La app no tiene autorización local. Generá la solicitud y pedí al propietario que comparta el Sheet con la cuenta de servicio y te devuelva el TOKEN.")


# ------------------------------------------------------------------------
# Preparar campos (materia, año, unidad)
# ------------------------------------------------------------------------
subjects, years, units = extract_candidates_from_sheets(sheets_dict)

# ------------------------------------------------------------------------
# Avatar: subir o presets (acepta GIFs)
# ------------------------------------------------------------------------
st.sidebar.markdown("---")
st.sidebar.markdown("### 👩‍🏫 Avatar del Tutor")
avatar_mode = st.sidebar.radio("Avatar", options=["Subir imagen/GIF", "Preset 1", "Preset 2", "Preset 3"], index=0)
avatar_bytes = None
if avatar_mode == "Subir imagen/GIF":
    avatar_upload = st.sidebar.file_uploader("Subí imagen o GIF", type=["png", "jpg", "jpeg", "gif"])
    if avatar_upload:
        avatar_bytes = avatar_upload.read()
else:
    # presets simples como SVG bytes
    svgs = {
        "Preset 1": """<svg viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg"> <circle cx="60" cy="60" r="58" fill="#3b82f6"/> <circle cx="60" cy="46" r="18" fill="#fff3"/> <circle cx="60" cy="72" r="22" fill="#ffedd5"/> </svg>""",
        "Preset 2": """<svg viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg"><rect rx="20" width="120" height="120" fill="#10b981"/><circle cx="60" cy="44" r="18" fill="#fff"/><ellipse cx="60" cy="80" rx="26" ry="18" fill="#fde68a"/></svg>""",
        "Preset 3": """<svg viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg"><circle cx="60" cy="60" r="58" fill="#7c3aed"/><circle cx="60" cy="46" r="18" fill="#fff"/><rect x="30" y="72" width="60" height="20" rx="8" fill="#fce7f3"/></svg>"""
    }
    avatar_bytes = svgs.get(avatar_mode).encode("utf-8")

# ------------------------------------------------------------------------
# Formulario de búsqueda (usa st.form para evitar duplicados)
# ------------------------------------------------------------------------
st.header("📚 Tutor IA — Consulta por materia / año / unidad")

with st.form("consulta_form", clear_on_submit=False):
    col1, col2, col3 = st.columns(3)
    with col1:
        subj_sel = st.selectbox("Materia (opcional)", options=["(no seleccionar)"] + (subjects or ["(no disponible)"]), key="form_subj")
    with col2:
        year_sel = st.selectbox("Año / Curso (opcional)", options=["(no seleccionar)"] + (years or ["(no disponible)"]), key="form_year")
    with col3:
        unit_sel = st.selectbox("Unidad / Tema (opcional)", options=["(no seleccionar)"] + (units or ["(no disponible)"]), key="form_unit")

    query = st.text_input("Tu pregunta o consulta (ej: 'recursos para fracciones en 3er año')", value="", key="form_query")
    submit = st.form_submit_button("Consultar ahora", help="Buscar contenidos y sugerencias prácticas")

# ------------------------------------------------------------------------
# Procesar la consulta
# ------------------------------------------------------------------------
def row_to_clean_text(row: pd.Series, max_chars=600):
    """Crea una vista compacta y sin fórmulas de las columnas relevantes de la fila."""
    parts = []
    for col in row.index:
        val = row[col]
        if val is None or (isinstance(val, str) and val.strip() == ""):
            continue
        # evitar mostrar fórmulas (ya sanitizado) y truncar textos largos
        text = str(val)
        if text.strip().startswith("="):
            continue
        parts.append(f"{col}: {text}")
    s = " | ".join(parts)
    if len(s) > max_chars:
        return s[:max_chars] + "…"
    return s

if submit:
    if not sheets_dict:
        st.warning("No hay datos cargados. Subí la credencial JSON o coloca el archivo Excel dentro de /data.")
    else:
        q = str(query or "").strip().lower()
        qwords = [w for w in q.split() if w]
        results = []
        # búsqueda heurística: filas que contienen las palabras, y que respeten filtros
        for sheet_name, df in sheets_dict.items():
            df_tmp = df.copy()
            for idx, row in df_tmp.iterrows():
                # construir text only from row (sanitized)
                row_text = " ".join([str(v).lower() for v in row.values if v is not None and str(v).strip() != ""])
                # aplicar filtros
                ok = True
                if subj_sel and subj_sel != "(no seleccionar)" and subj_sel.lower() not in row_text:
                    ok = False
                if year_sel and year_sel != "(no seleccionar)" and year_sel.lower() not in row_text:
                    ok = False
                if unit_sel and unit_sel != "(no seleccionar)" and unit_sel.lower() not in row_text:
                    ok = False
                if not ok:
                    continue
                score = sum(1 for w in qwords if w and w in row_text) if qwords else 0
                # si no hay query de texto y hay filtros, considerarlo con score 1
                if not qwords and (subj_sel != "(no seleccionar)" or year_sel != "(no seleccionar)" or unit_sel != "(no seleccionar)"):
                    score = max(score, 1)
                # guardar si tiene coincidencia o si no se buscó texto (mostrar ejemplos)
                if score > 0 or not qwords:
                    results.append({"score": score, "sheet": sheet_name, "row": int(idx), "row_obj": row})
        # ordenar por score
        results = sorted(results, key=lambda x: -x["score"])[:30]

        if not results:
            st.info("No encontré coincidencias. Probá otra búsqueda o dejá la consulta en blanco para ver ejemplos.")
        else:
            st.success(f"Encontré {len(results)} coincidencias. Mostrando top {min(10,len(results))}:")
            for i, r in enumerate(results[:10], start=1):
                st.markdown(f"**[{i}] Sheet:** {r['sheet']} — fila {r['row']} — puntuación={r['score']}")
                st.write(row_to_clean_text(r["row_obj"]))
                cols = st.columns([1,3])
                with cols[0]:
                    if st.button("Ver detalle", key=f"ver_detalle_{i}"):
                        st.write(r["row_obj"].to_dict())
                with cols[1]:
                    # boton para marcar como recurso útil (ej.)
                    if st.button("Marcar como Recurso útil", key=f"marcar_util_{i}"):
                        st.success("Marcado como recurso útil (no persistente).")
                st.markdown("---")

        # Si hay OpenAI key, generar sugerencias resumidas (opcional)
        if OPENAI_KEY and results:
            try:
                import openai
                openai.api_key = OPENAI_KEY
                context = "\n\n".join([f"Fuente: {r['sheet']}#{r['row']} -> " + row_to_clean_text(r['row_obj'], max_chars=500) for r in results[:4]])
                prompt = f"Contexto:\n{context}\n\nPregunta: {query}\n\nProponé 3 sugerencias prácticas para un docente (breve cada una)."
                with st.spinner("Generando sugerencias con LLM..."):
                    resp = openai.ChatCompletion.create(
                        model="gpt-3.5-turbo",
                        messages=[{"role":"user","content":prompt}],
                        max_tokens=400,
                        temperature=0.2
                    )
                    generated = resp.choices[0].message.content.strip()
                    st.subheader("Sugerencias generadas (OpenAI)")
                    st.write(generated)
                    st.session_state["last_generated"] = generated

                    if st.button("Guardar sugerencia en hoja 'Consejos'", key="guardar_consejo_1"):
                        try:
                            excels = list(DATA_DIR.glob("*.xlsx"))
                            if not excels:
                                st.error("No hay Excel local en /data para guardar. Exportá el Google Sheet primero.")
                            else:
                                path = excels[0]
                                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                                tmp.close()
                                wb = openpyxl.load_workbook(path)
                                if "Consejos" not in wb.sheetnames:
                                    wb.create_sheet("Consejos")
                                ws = wb["Consejos"]
                                srcs = " | ".join([f"{r['sheet']}#{r['row']}" for r in results[:4]])
                                ws.append([datetime.datetime.now().isoformat(), query, generated, srcs])
                                wb.save(tmp.name)
                                os.replace(tmp.name, str(path))
                                st.success("Sugerencia guardada en 'Consejos' ✅")
                        except Exception as e:
                            st.error("Fallo al guardar: " + str(e))
            except Exception as e:
                st.error("No se pudo usar OpenAI: " + str(e))

# ------------------------------------------------------------------------
# Panel lateral: mostrar avatar y TTS (Web Speech API)
# ------------------------------------------------------------------------
st.sidebar.markdown("---")
st.sidebar.markdown("### Avatar y audio")

if avatar_bytes:
    show_avatar_bytes(avatar_bytes, width=140)
else:
    st.sidebar.info("Subí una imagen/GIF arriba o elegí un preset.")

if st.sidebar.button("Escuchar última sugerencia (navegador)", key="tts_play"):
    last = st.session_state.get("last_generated", "")
    if last:
        js = f"""
        const msg = new SpeechSynthesisUtterance({json.dumps(last)});
        msg.rate = 0.95;
        window.speechSynthesis.speak(msg);
        """
        st.components.v1.html(f"<script>{js}</script>")
    else:
        st.sidebar.info("No hay texto generado para reproducir.")

# ------------------------------------------------------------------------
# Footer / ayuda
# ------------------------------------------------------------------------
st.markdown("---")
st.info("Sugerencia: si el propietario del Google Sheet no permite el acceso, pedí que comparta la hoja con la service account (el email aparece en la clave JSON).")
st.caption("No mostramos fórmulas del Excel. Si necesitás incluir fórmulas en los resultados, avisame para cambiar la política de visualización.")
