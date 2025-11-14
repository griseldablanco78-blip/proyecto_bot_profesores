import streamlit as st
import pandas as pd
import os

# --- CONFIGURACIÓN ---
st.set_page_config(page_title="Tutor IA para Profesores", page_icon="🎓", layout="wide")

st.title("🎓 Tutor IA para Profesores")
st.write("Bienvenido/a. Este asistente puede ayudarte a consultar información del ecosistema educativo y darte consejos personalizados.")

# Ruta al archivo Excel
excel_path = os.path.join("data", "Base de Datos Ecosistema Secundaria Aprende.xlsx")

# --- CARGA DEL EXCEL ---
if os.path.exists(excel_path):
    try:
        xls = pd.ExcelFile(excel_path)
        st.sidebar.success("📘 Archivo Excel cargado correctamente.")
        st.sidebar.write(f"Hojas detectadas: {', '.join(xls.sheet_names)}")
    except Exception as e:
        st.sidebar.error(f"Error al leer el Excel: {e}")
else:
    st.sidebar.warning("⚠️ No se encontró el archivo Excel en la carpeta 'data'.")

# --- CHAT SIMULADO ---
# --- Inserta / reemplaza la sección de Modo guiado / resultados por este bloque ---
import re
import openpyxl
import tempfile
from pathlib import Path
import base64
from io import BytesIO

import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_dataframe import get_as_dataframe

st.sidebar.markdown("## Cargar Google Sheet (opcional)")
uploaded_sa = st.sidebar.file_uploader("Credencial Service Account (JSON)", type=["json"])
sheet_url = st.sidebar.text_input("Pegar aquí la URL del Google Sheet (opcional)")

def get_gspread_client_from_file(json_path):
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(json_path, scope)
    return gspread.authorize(creds)

sheets_dict = {}
if uploaded_sa and sheet_url:
    sa_path = "data/sa_key.json"
    with open(sa_path, "wb") as f:
        f.write(uploaded_sa.getbuffer())
    try:
        client = get_gspread_client_from_file(sa_path)
        sh = client.open_by_url(sheet_url)
        for ws in sh.worksheets():
            df = get_as_dataframe(ws, evaluate_formulas=True)
            sheets_dict[ws.title] = df.fillna("")
        st.sidebar.success(f"Cargadas {len(sheets_dict)} hojas desde Google Sheets.")
    except Exception as e:
        st.sidebar.error("Error al leer el Google Sheet: " + str(e))
elif sheet_url and not uploaded_sa:
    st.sidebar.info("Pega la URL y subí la credencial JSON (service account) para que la app pueda leer el Sheet.")



# ---------- Avatars: subir o elegir preset ----------
st.sidebar.markdown("### 👩‍🏫 Avatar del Tutor")
avatar_choice = st.sidebar.radio("Elegir avatar", options=["Subir imagen","Avatar 1","Avatar 2","Avatar 3","Avatar 4"])

avatar_bytes = None
if avatar_choice == "Subir imagen":
    upload = st.sidebar.file_uploader("Subí una imagen (.png/.jpg). Opcional", type=["png","jpg","jpeg"])
    if upload:
        avatar_bytes = upload.read()
else:
    # avatares SVG sencillos embebidos (pequeños, inline)
    svgs = {
        "Avatar 1": """<svg viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg"> <circle cx="60" cy="60" r="58" fill="#3b82f6"/> <circle cx="60" cy="46" r="18" fill="#fff3"/> <circle cx="60" cy="72" r="22" fill="#ffedd5"/> </svg>""",
        "Avatar 2": """<svg viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg"><rect rx="20" width="120" height="120" fill="#10b981"/><circle cx="60" cy="44" r="18" fill="#fff"/><ellipse cx="60" cy="80" rx="26" ry="18" fill="#fde68a"/></svg>""",
        "Avatar 3": """<svg viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg"><rect width="120" height="120" rx="24" fill="#f97316"/><circle cx="60" cy="44" r="18" fill="#fff"/><circle cx="60" cy="78" r="20" fill="#fde68a"/></svg>""",
        "Avatar 4": """<svg viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg"><circle cx="60" cy="60" r="58" fill="#7c3aed"/><circle cx="60" cy="46" r="18" fill="#fff"/><rect x="30" y="72" width="60" height="20" rx="8" fill="#fce7f3"/></svg>"""
    }
    svg = svgs.get(avatar_choice)
    avatar_bytes = svg.encode("utf-8")

# función útil para mostrar bytes/SVG en Streamlit
def show_avatar(bts, width=120):
    if not bts:
        return
    try:
        # SVG?
        txt = bts.decode('utf-8')
        if txt.lstrip().startswith("<svg"):
            st.image("data:image/svg+xml;utf8," + txt, width=width)
            return
    except Exception:
        pass
    st.image(bts, width=width)


# ---------- Imports necesarios ----------
import os, json, datetime, tempfile
from pathlib import Path
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_dataframe import get_as_dataframe
import streamlit as st
from io import BytesIO
import openpyxl

# ---------- Sidebar: carga de Google Sheet y credencial + solicitante ----------
st.sidebar.title("Conectar Google Sheet (opcional)")
sheet_url = st.sidebar.text_input("Pegar URL del Google Sheet (opcional)")
requester_email = st.sidebar.text_input("Tu correo (para solicitar acceso si hace falta)")

uploaded_sa = st.sidebar.file_uploader("Credencial Service Account (JSON) — opcional", type=["json"])

# función para autorizar gspread desde archivo guardado
def get_gspread_client_from_file(json_path):
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(json_path, scope)
    return gspread.authorize(creds)

# función que devuelve dict of dataframes o raise error
def load_sheets_from_google(sheet_url, sa_path):
    client = get_gspread_client_from_file(sa_path)
    sh = client.open_by_url(sheet_url)
    sheets = {}
    for ws in sh.worksheets():
        df = get_as_dataframe(ws, evaluate_formulas=True)
        sheets[ws.title] = df.fillna("")
    return sheets

# Guardar JSON subido si existe
sa_path = None
if uploaded_sa:
    os.makedirs("data", exist_ok=True)
    sa_path = os.path.join("data", "sa_key.json")
    with open(sa_path, "wb") as f:
        f.write(uploaded_sa.getbuffer())

sheets_dict = {}
permission_ok = False
if sheet_url:
    if sa_path:
        try:
            sheets_dict = load_sheets_from_google(sheet_url, sa_path)
            permission_ok = True
            st.sidebar.success(f"Cargadas {len(sheets_dict)} hojas")
        except Exception as e:
            # error probable: permisos (ModuleNotFound handled elsewhere)
            st.sidebar.error("No fue posible leer el Sheet con esa credencial: " + str(e))
            # intentar extraer owner email del mensaje o mostrar plantilla
            permission_ok = False
    else:
        st.sidebar.info("Pega la URL y subí la credencial JSON (service account) para que la app pueda leer el Sheet.")

# Si no hay Google Sheet o no se cargó, permitir cargar Excel local
if not sheets_dict:
    # buscar excel local en data/
    data_folder = Path("data")
    local_excels = list(data_folder.glob("*.xlsx"))
    if local_excels:
        st.sidebar.success(f"Encontrado Excel local: {local_excels[0].name}", icon="💾")
        # ofrecemos opción para usarlo
        use_local = st.sidebar.checkbox("Usar Excel local en data/ (en lugar de Google Sheet)", value=True)
        if use_local:
            excel_path = str(local_excels[0])
            try:
                sheets_dict = pd.read_excel(excel_path, sheet_name=None)
                # fillna
                sheets_dict = {k: v.fillna("") for k, v in sheets_dict.items()}
            except Exception as e:
                st.sidebar.error("Error leyendo el Excel local: " + str(e))

# ---------- Si no tenemos acceso al sheet: mostrar plantilla de solicitud ----------
if sheet_url and not permission_ok:
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ¿No tenés acceso al Sheet?")
    owner_info = ("propietario@dominio.edu.ar" if requester_email else "PROPIETARIO_DEL_SHEET")
    # mostrar plantilla de correo editable
    subject = "Solicitud de acceso al Google Sheet para Tutor IA"
    body = f"""Hola,

Solicito por favor que compartan el Google Sheet con mi cuenta para usar la aplicación 'Tutor IA'.

URL: {sheet_url}
Correo solicitante: {requester_email or 'tu_correo@ejemplo.com'}

Por favor, compartan el archivo con la cuenta de servicio: <AQUÍ_SE_MUESTRA_EL_EMAIL_DE_LA_SERVICE_ACCOUNT> (subir la credencial JSON para ver el email) o con el correo de la app.
Gracias.
"""
    st.sidebar.markdown("Copiá y pegá este mensaje para pedir acceso:")
    st.sidebar.code(f"Subject: {subject}\n\n{body}", language="text")

# ---------- Inferir campos (materia, año, unidad) a partir de los sheets cargados ----------
# Generar una lista única de valores posibles para materia/año/unidad buscando columnas comunes
def extract_candidates_from_sheets(sheets):
    subjects, years, units = set(), set(), set()
    for name, df in sheets.items():
        for col in df.columns:
            cname = str(col).lower()
            if any(k in cname for k in ["materia","asign","subject","disciplina","curso","materi"]):
                subjects.update(df[col].dropna().astype(str).unique().tolist())
            if any(k in cname for k in ["año","anio","year","curso","grado","nivel"]):
                years.update(df[col].dropna().astype(str).unique().tolist())
            if any(k in cname for k in ["unidad","unidad temática","unidad_t","tema","capítulo","capitulo"]):
                units.update(df[col].dropna().astype(str).unique().tolist())
    return sorted([s for s in subjects if s]), sorted([y for y in years if y]), sorted([u for u in units if u])

subjects, years, units = extract_candidates_from_sheets(sheets_dict)



# ----------------- Formulario de consulta----------------------------------------------------
st.header("📚 Tutor IA — Consulta por materia / año / unidad")

# Usamos un form para evitar ids duplicados y controlar el submit
with st.form("consulta_form"):
    # muestra los selectboxes (usar variables subjects, years, units ya calculadas antes)
    col1, col2, col3 = st.columns(3)
    with col1:
        subj_sel = st.selectbox("Materia (opcional)", options=["(no seleccionar)"] + (subjects or ["(no disponible)"]), key="form_subj")
    with col2:
        year_sel = st.selectbox("Año / Curso (opcional)", options=["(no seleccionar)"] + (years or ["(no disponible)"]), key="form_year")
    with col3:
        unit_sel = st.selectbox("Unidad / Tema (opcional)", options=["(no seleccionar)"] + (units or ["(no disponible)"]), key="form_unit")

    query = st.text_input("Tu pregunta o consulta (puedes usar plantillas en Modo guiado)", value="", key="form_query")

    # Botón de envío del form (único submit)
    submit = st.form_submit_button("Consultar ahora", help="Enviar la consulta para buscar en el Excel / Sheets")

# Procesar la consulta cuando el usuario hace submit
if submit:
    if not sheets_dict:
        st.warning("No hay datos cargados. Subí la credencial JSON o coloca el archivo Excel dentro de /data.")
    else:
        qwords = [w.lower() for w in str(query).split() if w.strip()]
        hits = []
        # iterar sheets y buscar coincidencias (misma lógica que antes)
        for sheet_name, df in sheets_dict.items():
            df_tmp = df.copy()
            # forzar strings para evitar errores
            for col in df_tmp.columns:
                try:
                    df_tmp[col] = df_tmp[col].astype(str)
                except Exception:
                    pass
            for idx, row in df_tmp.iterrows():
                text = " ".join([str(v) for v in row.values]).lower()
                keep = True
                if subj_sel and subj_sel != "(no seleccionar)":
                    if subj_sel.lower() not in text:
                        keep = False
                if year_sel and year_sel != "(no seleccionar)":
                    if year_sel.lower() not in text:
                        keep = False
                if unit_sel and unit_sel != "(no seleccionar)":
                    if unit_sel.lower() not in text:
                        keep = False
                if not keep:
                    continue
                score = sum(1 for w in qwords if w and w in text)
                if score > 0 or not qwords:
                    hits.append((score, sheet_name, int(idx), row))
        hits = sorted(hits, key=lambda x: -x[0])[:50]

        if not hits:
            st.info("No encontré coincidencias exactas. Probá otra búsqueda o deja la consulta en blanco para ver ejemplos.")
        else:
            st.success(f"Encontré {len(hits)} coincidencias. Mostrando top {min(10,len(hits))}:")
            for i, (score, sheet_name, idx, row) in enumerate(hits[:10], start=1):
                st.markdown(f"**[{i}] Sheet:** {sheet_name} — fila {idx} — puntuación={score}")
                preview = []
                for col in row.index:
                    preview.append(f"**{col}**: {row[col]}")
                st.write("\n".join(preview))
                # Si querés botones por cada hit, agregalos con keys únicos:
                if st.button("Ver detalle", key=f"ver_detalle_{i}"):
                    st.write(row.to_dict())
                st.markdown("---")

        # Si hay OpenAI disponible, generar sugerencias (opcional)
        OPENAI_KEY = os.getenv("OPENAI_API_KEY")
        if OPENAI_KEY and hits:
            try:
                import openai
                openai.api_key = OPENAI_KEY
                context = "\n\n".join([f"Fuente: sheet={h[1]}, row={h[2]}:\n" + " ".join([str(v) for v in h[3].values]) for h in hits[:4]])
                prompt = f"Contexto:\n{context}\n\nPregunta: {query}\n\nProponé 3 sugerencias pedagógicas breves."
                resp = openai.ChatCompletion.create(
                    model="gpt-3.5-turbo",
                    messages=[{"role":"user","content":prompt}],
                    max_tokens=400, temperature=0.2
                )
                generated = resp.choices[0].message.content.strip()
                st.subheader("Respuesta generada (OpenAI)")
                st.write(generated)
                # guardamos en session_state para TTS o futura acción
                st.session_state["last_generated"] = generated

                # Botón para guardar: dar key único
                if st.button("Guardar esta respuesta como Consejo en Excel", key="guardar_consejo_form"):
                    try:
                        excels = list(Path("data").glob("*.xlsx"))
                        if not excels:
                            st.error("No hay Excel local en /data para guardar. Exportá el Google Sheet primero.")
                        else:
                            path = str(excels[0])
                            wb = openpyxl.load_workbook(path)
                            if "Consejos" not in wb.sheetnames:
                                wb.create_sheet("Consejos")
                            ws = wb["Consejos"]
                            sources = " | ".join([f"{h[1]}#{h[2]}" for h in hits[:4]])
                            ws.append([datetime.datetime.now().isoformat(), requester_email or "manual", query, generated, sources])
                            wb.save(path)
                            st.success("Guardado en hoja 'Consejos' del Excel local ✅")
                    except Exception as e:
                        st.error("Fallo al guardar en Excel: " + str(e))
            except Exception as e:
                st.error("No se pudo usar OpenAI: " + str(e))



# ---------- Avatar: mostrar imagen estática o GIF y control TTS (navegador) ----------
st.sidebar.markdown("---")
st.sidebar.markdown("### Avatar del Tutor")
avatar_file = st.sidebar.file_uploader("Subí una imagen o GIF (.png/.jpg/.gif)", type=["png","jpg","jpeg","gif"])
avatar_choice = st.sidebar.radio("O elegir uno de los presets", options=["(ninguno)","Avatar 1","Avatar 2","Avatar 3"], index=0)

def show_avatar_from_bytes(bts):
    if not bts: return
    st.sidebar.image(bts, width=140)

if avatar_file:
    show_avatar_from_bytes(avatar_file.read())
else:
    # presets (simple colored svgs o una imagen local)
    if avatar_choice != "(ninguno)":
        st.sidebar.image(f"assets/{avatar_choice}.png", width=140)  # prepara assets

# TTS: boton que usa Web Speech API (en el navegador) -> generamos JS para reproducir texto
def tts_js(text):
    js = f"""
    const msg = new SpeechSynthesisUtterance({json.dumps(text)});
    msg.rate = 0.95;
    window.speechSynthesis.speak(msg);
    """
    return js

if st.button("Escuchar la última respuesta (navegador TTS)"):
    # recolectar último texto generado si existe
    last = st.session_state.get("last_generated", None)
    if last:
        st.components.v1.html(f"<script>{tts_js(last)}</script>")
    else:
        st.info("Primero generá una respuesta con OpenAI para poder reproducirla.")


# ---------- Modo guiado y templates ----------
st.header("Chat del Tutor — Modo guiado y guardado")
if 'df_dict' not in locals():
    # intentar cargar automáticamente (si ya lo hiciste antes)
    data_folder = Path(os.getcwd()) / "data"
    files = list(data_folder.glob("*.xlsx"))
    df_dict = {}
    if files:
        try:
            df_dict = pd.read_excel(files[0], sheet_name=None)
        except Exception as e:
            st.error(f"No pude leer el Excel: {e}")

# extraer subjects/years si existen
possible_subjects, possible_years = set(), set()
for sheet, df in (df_dict.items() if df_dict else []):
    for c in df.columns:
        if re.search(r"materi|asign|discipl", c, re.I):
            possible_subjects.update(df[c].dropna().astype(str).unique().tolist())
        if re.search(r"año|anio|year|curso|grado", c, re.I):
            possible_years.update(df[c].dropna().astype(str).unique().tolist())
possible_subjects = sorted([s for s in possible_subjects if s])[:200]
possible_years = sorted([y for y in possible_years if y])[:200]

with st.expander("Modo guiado (plantillas)"):
    guided = st.checkbox("Activar modo guiado", value=True)
    subj_choice = "(no usar)"
    year_choice = "(no usar)"
    if possible_subjects:
        subj_choice = st.selectbox("Materia (para plantillas)", options=["(no usar)"] + possible_subjects)
    else:
        subj_choice = st.text_input("Materia (opcional)")
    if possible_years:
        year_choice = st.selectbox("Año/curso (para plantillas)", options=["(no usar)"] + possible_years)
    else:
        year_choice = st.text_input("Año/Curso (opcional)")
    topic_input = st.text_input("Tema / tópico (ej: fracciones, comprensión lectora)")

    templates = [
        "¿Qué contenidos recomienda la materia {subject} para el {year} año?",
        "Proponé 3 actividades prácticas para trabajar {topic} en {subject} ({year} año).",
        "¿Qué evaluación sugerida hay para {subject} en {year} año?",
        "¿Qué competencias se trabajan en {subject} relacionadas con {topic}?",
        "¿Qué recursos digitales se sugieren para enseñar {topic} en {subject}?",
        "Diseñá una secuencia de 3 clases sobre {topic} para {year} año en {subject}.",
        "Proponé una actividad interdisciplinaria entre {subject} y otra materia para trabajar {topic}.",
        "¿Cómo adaptar la propuesta para estudiantes con dificultades de aprendizaje en {topic}?"
    ]

    cols = st.columns(2)
    for i, t in enumerate(templates):
        filled = t.format(
            subject=(subj_choice if subj_choice and subj_choice != "(no usar)" else "{subject}"),
            year=(year_choice if year_choice and year_choice != "(no usar)" else "{year}"),
            topic=(topic_input if topic_input else "{topic}")
        )
        with cols[i % 2]:
            if st.button(filled[:80], key=f"templ_{i}"):
                st.session_state['guided_question'] = filled

if 'guided_question' not in st.session_state:
    st.session_state['guided_question'] = ""

# input y mostrar avatar al lado
left, right = st.columns([4,1])
with left:
    user_query = st.text_input("Tu pregunta (o usá una plantilla):", value=st.session_state['guided_question'])
    if st.button("Consultar ahora", key="consultar_ahora_1"):
        q = user_query.strip()
        if not q:
            st.info("Escribí o elegí una pregunta.")
        else:
            # recuperación semántica si existe índice, si no búsqueda simple
            index_exists = os.path.exists("index/faiss.index") and os.path.exists("index/metadata.json")
            contexts = []
            if index_exists:
                try:
                    contexts = retrieve(q, top_k=6, sheet_filter=None)
                except Exception:
                    # fallback a búsqueda básica en metadata.json
                    with open("index/metadata.json","r",encoding="utf-8") as f:
                        meta = json.load(f)
                    hits=[]
                    qwords=q.lower().split()
                    for i,item in enumerate(meta):
                        txt=item.get("text","").lower()
                        score=sum(1 for w in qwords if w and w in txt)
                        if score>0:
                            hits.append((score,i,item))
                    hits=sorted(hits,key=lambda x:-x[0])[:6]
                    contexts=[h[2] for h in hits]
            else:
                # búsqueda por contenido en df_dict
                qwords=[w.lower() for w in re.findall(r'\w+', q)]
                hits=[]
                for sheet, df in (df_dict.items() if df_dict else []):
                    for idx, row in df.fillna("").iterrows():
                        text=" ".join([str(v) for v in row.values])
                        score=sum(1 for w in qwords if w and w in text.lower())
                        if score>0:
                            hits.append((score, sheet, idx, text))
                hits=sorted(hits,key=lambda x:-x[0])[:10]
                contexts=[{"metadata":{"sheet":h[1],"row_index":h[2]},"text":h[3]} for h in hits]

            if not contexts:
                st.warning("No encontré resultados relevantes.")
            else:
                # mostrar avatar y la primera respuesta posible
                st.subheader("Fuentes encontradas")
                for i,c in enumerate(contexts,1):
                    st.markdown(f"**[{i}] sheet={c['metadata'].get('sheet')} row={c['metadata'].get('row_index')}**")
                    st.write(c['text'][:800])
                    st.markdown("---")

                # Si tenés OpenAI, generar texto sugerido
                suggested_answer = None
                if OPENAI_KEY:
                    try:
                        import openai
                        openai.api_key = OPENAI_KEY
                        system = "Eres un asistente pedagógico que sugiere actividades y mejoras breves."
                        context_text = "\n\n".join([f"Fuente (sheet={c['metadata']['sheet']}, row={c['metadata']['row_index']}):\n{c['text']}" for c in contexts])
                        prompt = f"Contexto:\n{context_text}\n\nPregunta: {q}\n\nProponé 3 alternativas prácticas y breves para un docente."
                        with st.spinner("Generando respuesta con LLM..."):
                            resp = openai.ChatCompletion.create(
                                model="gpt-3.5-turbo",
                                messages=[{"role":"system","content":system},{"role":"user","content":prompt}],
                                max_tokens=450,
                                temperature=0.3
                            )
                            suggested_answer = resp.choices[0].message.content.strip()
                    except Exception as e:
                        st.error("No se pudo generar respuesta LLM: " + str(e))

                # mostrar avatar a la derecha del resultado
                if suggested_answer:
                    # layout con avatar
                    c1, c2 = st.columns([6,1])
                    with c1:
                        st.subheader("Respuesta sugerida")
                        st.write(suggested_answer)
                    with c2:
                        show_avatar(avatar_bytes, width=120)

                    # botón para guardar esta respuesta como fila en la hoja 'Consejos'
                    if st.button("Guardar esta respuesta en la hoja 'Consejos'"):
                        # seguridad: hacer backup temporal y append
                        try:
                            data_folder = Path(os.getcwd()) / "data"
                            files = list(data_folder.glob("*.xlsx"))
                            if not files:
                                st.error("No encontré el archivo Excel para guardar. Asegurate que exista en /data.")
                            else:
                                excel_path = files[0]
                                # hacer copia temporal
                                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                                tmp.close()
                                wb = openpyxl.load_workbook(excel_path)
                                sheet_name = "Consejos"
                                if sheet_name not in wb.sheetnames:
                                    wb.create_sheet(sheet_name)
                                ws = wb[sheet_name]
                                # fila: Timestamp | Pregunta | Respuesta | FuenteSheets (concatenado)
                                srcs = " | ".join([f"{c['metadata'].get('sheet')}#{c['metadata'].get('row_index')}" for c in contexts])
                                ws.append([datetime.datetime.now().isoformat(), q, suggested_answer, srcs])
                                wb.save(tmp.name)
                                # reemplazar archivo original (más seguro que escribir directo sobre OneDrive)
                                os.replace(tmp.name, str(excel_path))
                                st.success("Respuesta guardada en la hoja 'Consejos' ✅")
                        except Exception as e:
                            st.error("Error al guardar en Excel: " + str(e))
                else:
                    # si no hay LLM, mostrar avatar y opción de guardar uno de los textos mostrados
                    show_avatar(avatar_bytes, width=120)
                    st.info("Para guardar una de las fuentes mostradas, seleccioná el número y presioná Guardar abajo.")
                    sel = st.number_input("Número de fuente a guardar (1..)", min_value=1, max_value=len(contexts), value=1)
                    if st.button("Guardar fuente seleccionada en hoja 'Consejos'"):
                        chosen = contexts[sel-1]
                        try:
                            data_folder = Path(os.getcwd()) / "data"
                            files = list(data_folder.glob("*.xlsx"))
                            if not files:
                                st.error("No encontré el archivo Excel para guardar.")
                            else:
                                excel_path = files[0]
                                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                                tmp.close()
                                wb = openpyxl.load_workbook(excel_path)
                                sheet_name = "Consejos"
                                if sheet_name not in wb.sheetnames:
                                    wb.create_sheet(sheet_name)
                                ws = wb[sheet_name]
                                srcs = f"{chosen['metadata'].get('sheet')}#{chosen['metadata'].get('row_index')}"
                                ws.append([datetime.datetime.now().isoformat(), q, chosen['text'], srcs])
                                wb.save(tmp.name)
                                os.replace(tmp.name, str(excel_path))
                                st.success("Fuente guardada en la hoja 'Consejos' ✅")
                        except Exception as e:
                            st.error("Error al guardar en Excel: " + str(e))
with right:
    # panel pequeño con avatar
    st.markdown("**Avatar actual**")
    show_avatar(avatar_bytes, width=120)
